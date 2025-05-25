from  openpyxl import  *
import math

class Package():
    def __init__(self):
        pass

class Wallpaper(Package):
    def __init__(self):
        super().__init__()

class Calculation_Wallpaper(Wallpaper):
    def __init__(self):
        super().__init__()
    def  calculation_quantity(*args):
        wallpaper_strips = args[0] / (args[3] * .1)
        strips_in_one_roll = args[4] / args[2] 
        return  wallpaper_strips / strips_in_one_roll
    def  calculation_cost(*args):
        return  args[0] * args[1]

class Tile(Package):
    def __init__(self):
        super().__init__()

class Calculation_Tile(Tile):
    def __init__(self):
        super().__init__()
    def  T_calculation_length(*args):
        return math.ceil(args[0] * args[1] / (args[2] * 0.01) / (args[3] * 0.01))
    def  T_calculation_cost(*args):
        return  args[0] * args[1]

class Laminat(Package):
    def __init__(self):
        super().__init__()

class Calculation_Laminat(Laminat):
    def __init__(self):
        super().__init__()
    def  L_calculation_length(*args):
        S = args[0] * args[1]
        return  S / args[2]
    def  L_calculation_cost(*args):
        return  args[0] * args[1]

class Save_Wallpaper(Wallpaper):
    def __init__(self):
        super().__init__()
    def  data_Save(*args):
        wb=load_workbook("Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['A2'] = "Количество:"
            ws['B2'] = args[0]

            ws['A3'] = "Общая цена:"
            ws['B3'].value = args[1]
            wb.save("Result.xlsx")
        '''elif args[2] != 0:
            ws['C2'] = "Количество:"
            ws['D2'] = args[2]
            ws['C3'] = "Общая цена:"
            ws['D3'].value = args[3]
            wb.save("Result.xlsx")
        else:
            ws['E2'] = "Количество:"
            ws['F2'] = args[4]
            ws['E3'] = "Общая цена:"
            ws['F3'].value = args[5]
            wb.save("Result.xlsx")'''

class Save_Tile(Tile):
    def __init__(self):
        super().__init__()
    def  data_Save(*args):
        wb=load_workbook("Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['C2'] = "Количество:"
            ws['D2'] = args[0]

            ws['C3'] = "Общая цена:"
            ws['D3'].value = args[1]
            wb.save("Result.xlsx")

class Save_Laminat(Laminat):
    def __init__(self):
        super().__init__()
    def  data_Save(*args):
        wb=load_workbook("Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['E2'] = "Количество:"
            ws['F2'] = args[0]
            ws['E3'] = "Общая цена:"
            ws['F3'].value = args[1]
            wb.save("Result.xlsx")